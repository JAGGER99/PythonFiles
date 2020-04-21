# Excel Spread sheets Demo:

import openpyxl as xl # makes code shorter
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("trans.xlsx")
sheet = wb["Sheet1"]       # accesses the first sheet # must be capital! ---> "Sheet1"
  

for row in range(2, sheet.max_row + 1):
    
   cell = sheet.cell(row,3) #actual column 3
   corrected_price = cell.value*.9
   corrected_price_cell = sheet.cell(row, 4) #actual column 4
   corrected_price_cell.value = corrected_price
      
    ########################################################################
values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col =4, max_col = 4) #specifying what values you want to create a chart for
chart = BarChart()
chart.add_data(values) # adding the values to your chart
sheet.add_chart(chart, 'E2') #adding chart to excel sheet specifically at e2 cell
#wb.save("trans2.xlsx") # saving the changes
wb.save("trans2.xlsx")  


########################################################
########################################################
########################################################
    

#cell = sheet["a1"]          # accesses the first cell in sheet 1

#sheet.cell(1,1)             #another way of "accessing" (not displaying) the first cell in sheet 1
#print(cell.value)           # "displays" what is in the first cell

#print(sheet.max_row)        # command for finding how many rows you have in your spread sheet
