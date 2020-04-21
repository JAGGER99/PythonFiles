import openpyxl as xl

wb = xl.load_workbook("Combo_Info.xlsx")
sheet = wb["Sheet1"] 
####################################################

print("ENTREES BY PRICE: ")    
for row in range(2, sheet.max_row + 1):
    
   cell_entrees = sheet.cell(row,1)
   cell_entrees_prices = sheet.cell(row,2)  
   print(f"{cell_entrees.value}  ${cell_entrees_prices.value}")
          
####################################################

print()
print("SIDES BY PRICE: ")    
for row in range(2, sheet.max_row -2):
    
   cell_sides = sheet.cell(row,5)
   cell_side_prices = sheet.cell(row,6)
   print(f"{cell_sides.value}  ${cell_side_prices.value}")

print()
#####################################################

print("ENTREES BY CALORIES: ")    
for row in range(2, sheet.max_row + 1):
    
   cell_entrees = sheet.cell(row,1)
   cell_entrees_cals = sheet.cell(row,3)  
   print(f"{cell_entrees.value}  {cell_entrees_cals.value}")
          
print()

####################################################

print("SIDES BY CALORIES: ")    
for row in range(2, sheet.max_row -2):
    
   cell_sides = sheet.cell(row,5)
   cell_side_cals = sheet.cell(row,7)
   print(f"{cell_sides.value}  {cell_side_cals.value}")
   
print()
#######################################################
#######################################################
#######################################################
Entree_Ratio =[]
for row in range(2, sheet.max_row + 1):
    entree_ratio_cell = sheet.cell(row,4)
    Entree_Ratio.append(entree_ratio_cell.value)  

print(Entree_Ratio)
print()

Side_Ratio = []
for row in range(2, sheet.max_row-2):
    side_ratio_cell = sheet.cell(row,8)
    Side_Ratio.append(side_ratio_cell.value)  


print(Side_Ratio)
print()
#################################################


Combos = []
combo_max = 0
for x in range(10): #the range of the sides...10 (0-9) itmes
    for y in range(13): #the range pf the entrees....13 (0-12) items
        Combos.append(Entree_Ratio[y] + Side_Ratio[x])
        if Entree_Ratio[y] + Side_Ratio[x] > combo_max:
            combo_max = Entree_Ratio[y] + Side_Ratio[x]
            entree_max_index = y
            side_max_index = x

print(Combos)

Y= sheet.cell(entree_max_index+2,1) #+1 to the row for converting from zero based to 1 based and +1 because the first row is names!
X= sheet.cell(side_max_index+2,5) #+1 to the row for converting from zero based to 1 based and +1 because the first row is names!

print()
print("The Winning Combo!!!:")
print(f"Entree: {Y.value}")
print(f"Side: {X.value}")      

